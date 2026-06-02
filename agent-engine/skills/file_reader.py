#!/usr/bin/env python3
"""
File Reader Skill — Read and analyze files (PDF, images, code, documents)
"""

import os
import json
import base64
from pathlib import Path
from typing import Optional, Dict, Any


def read_file(file_path: str) -> Dict[str, Any]:
    """Read and analyze a file, returning structured content."""
    path = Path(file_path).expanduser().resolve()
    
    if not path.exists():
        return {"error": f"File not found: {file_path}"}
    
    if not path.is_file():
        return {"error": f"Not a file: {file_path}"}
    
    suffix = path.suffix.lower()
    size = path.stat().st_size
    
    result = {
        "path": str(path),
        "name": path.name,
        "size": size,
        "type": "unknown",
    }
    
    # Text files (code, config, markdown)
    if suffix in [".txt", ".md", ".py", ".js", ".ts", ".jsx", ".tsx", ".json", ".yaml", ".yml", ".toml", ".ini", ".cfg", ".sh", ".bash", ".zsh", ".fish", ".html", ".css", ".scss", ".less", ".xml", ".csv", ".log", ".env", ".gitignore", ".dockerignore"]:
        try:
            content = path.read_text(encoding="utf-8", errors="ignore")
            result["type"] = "text"
            result["content"] = content
            result["lines"] = content.count("\n") + 1
            
            # Language detection
            lang_map = {
                ".py": "python", ".js": "javascript", ".ts": "typescript",
                ".jsx": "jsx", ".tsx": "tsx", ".json": "json",
                ".yaml": "yaml", ".yml": "yaml", ".toml": "toml",
                ".sh": "bash", ".bash": "bash", ".zsh": "zsh",
                ".html": "html", ".css": "css", ".scss": "scss",
                ".xml": "xml", ".csv": "csv", ".md": "markdown",
            }
            result["language"] = lang_map.get(suffix, "text")
        except Exception as e:
            result["error"] = f"Failed to read text: {e}"
    
    # Images
    elif suffix in [".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp", ".svg"]:
        try:
            with open(path, "rb") as f:
                data = base64.b64encode(f.read()).decode("utf-8")
            result["type"] = "image"
            result["format"] = suffix[1:]
            result["base64"] = data
            result["data_url"] = f"data:image/{suffix[1:]};base64,{data}"
        except Exception as e:
            result["error"] = f"Failed to read image: {e}"
    
    # PDF
    elif suffix == ".pdf":
        try:
            import PyPDF2
            with open(path, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                pages = []
                for i, page in enumerate(reader.pages[:10]):  # Limit to first 10 pages
                    text = page.extract_text() or ""
                    pages.append({"page": i + 1, "text": text})
                result["type"] = "pdf"
                result["pages"] = pages
                result["total_pages"] = len(reader.pages)
        except ImportError:
            result["error"] = "PyPDF2 not installed. Run: pip install PyPDF2"
        except Exception as e:
            result["error"] = f"Failed to read PDF: {e}"
    
    # Office documents
    elif suffix in [".docx", ".doc"]:
        try:
            import docx
            doc = docx.Document(path)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            result["type"] = "document"
            result["format"] = "docx"
            result["content"] = "\n".join(paragraphs)
        except ImportError:
            result["error"] = "python-docx not installed. Run: pip install python-docx"
        except Exception as e:
            result["error"] = f"Failed to read document: {e}"
    
    elif suffix in [".xlsx", ".xls"]:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames[:5]:  # Limit to first 5 sheets
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(max_row=50, values_only=True):  # Limit to 50 rows
                    rows.append([str(cell) if cell is not None else "" for cell in row])
                sheets[sheet_name] = rows
            result["type"] = "spreadsheet"
            result["format"] = "xlsx"
            result["sheets"] = sheets
            result["total_sheets"] = len(wb.sheetnames)
        except ImportError:
            result["error"] = "openpyxl not installed. Run: pip install openpyxl"
        except Exception as e:
            result["error"] = f"Failed to read spreadsheet: {e}"
    
    # Archives
    elif suffix in [".zip", ".tar", ".gz", ".bz2", ".7z", ".rar"]:
        result["type"] = "archive"
        result["format"] = suffix[1:]
        result["note"] = "Archive file detected. Use terminal commands to extract."
    
    # Binary/unknown
    else:
        result["type"] = "binary"
        result["note"] = f"Binary file ({suffix}). Use terminal commands to process."
    
    return result


def list_directory(dir_path: str = ".", max_depth: int = 2) -> Dict[str, Any]:
    """List directory contents with structure."""
    path = Path(dir_path).expanduser().resolve()
    
    if not path.exists():
        return {"error": f"Directory not found: {dir_path}"}
    
    if not path.is_dir():
        return {"error": f"Not a directory: {dir_path}"}
    
    def scan(p: Path, depth: int = 0) -> Dict:
        if depth > max_depth:
            return None
        
        result = {"name": p.name, "type": "directory", "children": []}
        try:
            for item in sorted(p.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
                if item.name.startswith("."):
                    continue
                
                if item.is_dir():
                    child = scan(item, depth + 1)
                    if child:
                        result["children"].append(child)
                else:
                    result["children"].append({
                        "name": item.name,
                        "type": "file",
                        "size": item.stat().st_size,
                    })
        except PermissionError:
            result["error"] = "Permission denied"
        
        return result
    
    return scan(path)


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        result = read_file(sys.argv[1])
        print(json.dumps(result, indent=2))
    else:
        print("Usage: python file_reader.py <file_path>")
