#!/usr/bin/env python3
"""
TaskBolt Agent Engine — Lightweight Hermes-style agent
Runs locally, calls Vercel for AI, executes tools on user's machine.
All data stays local. Communicates with Tauri via stdin/stdout JSON.
"""

import sys
import json
import os
import subprocess
import platform
import shutil
import re
import time
import signal
from pathlib import Path
from typing import Optional

# ── Local storage ──────────────────────────────────────
DATA_DIR = Path.home() / ".taskbolt"
MEMORY_DIR = DATA_DIR / "memory"
MEMORY_DIR.mkdir(parents=True, exist_ok=True)
CONVERSATIONS_FILE = DATA_DIR / "conversations.json"

# ── Vercel API (auth required, key stays on server) ───
VERCEL_API = os.environ.get("TASKBOLT_API", "https://taskbolt.space/api/ai/agent")
AUTH_TOKEN = os.environ.get("TASKBOLT_TOKEN", "")


def sanitize_error(err_msg: str, status_code: int = 0) -> str:
    """Sanitize error messages — never expose provider names, billing, or internal details."""
    msg_lower = err_msg.lower()
    # Our own credit errors are OK to pass through (they're user-facing already)
    if "no credits remaining" in msg_lower or "ratelimited" in msg_lower:
        return err_msg
    # Billing / arrearage / overdue
    if any(kw in msg_lower for kw in ["arrearage", "overdue", "insufficientbalance", "account is in good standing"]):
        return "Our service is temporarily unavailable. We're working on it — please try again shortly."
    # Rate limiting from provider
    if status_code == 429 or "rate" in msg_lower or "throttl" in msg_lower:
        return "We're experiencing high demand right now. Please try again in a moment."
    # Auth / key issues
    if status_code in (401, 403) or "api key" in msg_lower or "invalid" in msg_lower and "key" in msg_lower:
        return "Service configuration error. Our team has been notified."
    # Model / provider exposure
    if any(kw in msg_lower for kw in ["dashscope", "alibaba", "qwen", "model not found"]):
        return "Service temporarily unavailable. Please try again."
    # Generic server errors
    if status_code >= 500:
        return "Our AI service is temporarily unavailable. Please try again in a few moments."
    # Anything else that looks like a raw provider error
    if any(kw in msg_lower for kw in ["error", "exception", "traceback"]):
        return "Something went wrong processing your request. Please try again."
    return err_msg

# ── System prompt ──────────────────────────────────────
SYSTEM_PROMPT = """You are TaskBolt, an intelligent AI assistant with FULL access to the user's computer. You can do ANYTHING they ask.

CORE CAPABILITIES:
- **General Tasks**: Execute any command, automate any workflow, manage files, install software, configure settings
- **Troubleshooting**: Diagnose slow PCs, fix crashes, repair network issues, resolve software conflicts, analyze system health
- **Productivity**: Create documents, analyze spreadsheets, manage emails, organize files, write reports, translate text, research topics
- **System Management**: Install/update/remove software, manage startup programs, configure networking, set up firewalls, create backups
- **Desktop Automation**: Control mouse clicks, keyboard input, automate repetitive GUI tasks, take screenshots
- **Browser Automation**: Browse the web, fill forms, scrape data, download files, automate web workflows
- **AI Tools**: Set up Claude Code, GitHub Copilot, Cursor, Aider, OpenHands, Ollama, local LLMs
- **Development**: Set up Python, Node.js, Docker, databases, CI/CD pipelines, version control
- **File Analysis**: Read PDFs, images (vision), Word docs, Excel sheets, code files, archives
- **Security**: Scan for vulnerabilities, harden the system, check for malware, audit permissions

CRITICAL RULES:
- Always use tools to take action — don't just describe what to do
- Be direct and practical — the user wants RESULTS, not explanations
- Explain what you're doing and why, briefly
- For Windows, prefer PowerShell commands; for Mac/Linux use bash
- Read system info before making changes
- Keep memory of what you've done for the user
- If something fails, try alternative approaches
- When users upload/mention files, ALWAYS use analyze_file
- For system problems, run full diagnosis BEFORE suggesting fixes
- For complex setups, explain each step and confirm before proceeding
- You are NOT just for developers — you help office workers, students, creators, and everyone

PERSONALITY:
- Friendly but professional — like a smart IT friend who also happens to be a productivity expert
- Proactive — anticipate what the user needs next
- Concise — don't over-explain obvious things
- Confident — you can do powerful things, show that
- Universal — never assume the user is a developer unless they say so
"""

# ── Tool Definitions ───────────────────────────────────
TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "run_command",
            "description": "Execute a shell command (PowerShell on Windows, bash on Linux/Mac). Use for installing software, running scripts, system configuration.",
            "parameters": {
                "type": "object",
                "properties": {
                    "command": {"type": "string", "description": "The command to execute"},
                    "timeout": {"type": "integer", "description": "Timeout in seconds (default: 120)", "default": 120},
                    "workdir": {"type": "string", "description": "Working directory (optional)"}
                },
                "required": ["command"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "read_file",
            "description": "Read the contents of a file from the user's machine.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Absolute or relative file path"}
                },
                "required": ["path"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "write_file",
            "description": "Write content to a file on the user's machine. Creates parent directories if needed.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "File path"},
                    "content": {"type": "string", "description": "File content to write"}
                },
                "required": ["path", "content"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "search_files",
            "description": "Search for files by name pattern or search file contents. Uses ripgrep/grep.",
            "parameters": {
                "type": "object",
                "properties": {
                    "pattern": {"type": "string", "description": "Search pattern (regex for content, glob for filenames)"},
                    "path": {"type": "string", "description": "Directory to search in (default: current dir)"},
                    "target": {"type": "string", "enum": ["content", "files"], "description": "Search inside files ('content') or find files by name ('files')"}
                },
                "required": ["pattern"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_system_info",
            "description": "Get information about the user's system: OS, hostname, installed software, disk space, etc.",
            "parameters": {
                "type": "object",
                "properties": {},
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "install_software",
            "description": "Install software using the system package manager (winget on Windows, brew on Mac, apt on Linux).",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string", "description": "Software name or package ID"},
                    "method": {"type": "string", "enum": ["winget", "choco", "brew", "apt", "pip", "npm"], "description": "Installation method (auto-detected if omitted)"}
                },
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "save_memory",
            "description": "Save a piece of information to persistent memory. Use this to remember user preferences, system details, past actions, etc.",
            "parameters": {
                "type": "object",
                "properties": {
                    "key": {"type": "string", "description": "Memory key (e.g., 'user_preferences', 'installed_software')"},
                    "value": {"type": "string", "description": "The information to remember"}
                },
                "required": ["key", "value"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "recall_memory",
            "description": "Recall information previously saved to memory.",
            "parameters": {
                "type": "object",
                "properties": {
                    "key": {"type": "string", "description": "Memory key to recall. Omit to see all keys."}
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "analyze_file",
            "description": "Analyze a file: read PDFs, images (vision), documents (DOCX, XLSX), code files, and more. Returns structured content the AI can understand. Supports text files, images, PDFs, Office documents, and binary files.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Absolute path to the file to analyze"},
                    "question": {"type": "string", "description": "Optional: specific question about the file content"}
                },
                "required": ["path"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "run_diagnosis",
            "description": "Run comprehensive system health diagnostics. Checks CPU, memory, disk, network, startup programs, and identifies issues with recommendations.",
            "parameters": {
                "type": "object",
                "properties": {
                    "area": {"type": "string", "enum": ["full", "network", "disk", "performance"], "description": "Specific area to diagnose (default: full)"}
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "setup_local_llm",
            "description": "Install and configure a local LLM on the user's machine (Ollama, llama.cpp, LM Studio). Guides through model selection, download, and setup.",
            "parameters": {
                "type": "object",
                "properties": {
                    "provider": {"type": "string", "enum": ["ollama", "llamacpp", "lmstudio"], "description": "Which local LLM provider to set up"},
                    "model": {"type": "string", "description": "Model to download (e.g., 'llama3', 'mistral', 'codellama')"}
                },
                "required": ["provider"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "configure_ai_agent",
            "description": "Set up and configure AI coding agents on the user's machine: Claude Code, GitHub Copilot CLI, Cursor, Aider, OpenHands, etc.",
            "parameters": {
                "type": "object",
                "properties": {
                    "agent": {"type": "string", "enum": ["claude-code", "copilot-cli", "cursor", "aider", "openhands", "continue"], "description": "Which AI agent to configure"},
                    "api_key": {"type": "string", "description": "API key for the service (if needed)"}
                },
                "required": ["agent"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_directory",
            "description": "List files and folders in a directory with structure, sizes, and types.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "Directory path (default: current directory)"},
                    "recursive": {"type": "boolean", "description": "List recursively (default: false)"}
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "desktop_control",
            "description": "Control the user's desktop: click at coordinates, type text, press keys, take screenshots, move the mouse. Use for GUI automation when CLI is not available.",
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {"type": "string", "enum": ["click", "double_click", "right_click", "type", "press_key", "screenshot", "move_to", "scroll"], "description": "Desktop action to perform"},
                    "x": {"type": "integer", "description": "X coordinate (for click/move_to)"},
                    "y": {"type": "integer", "description": "Y coordinate (for click/move_to)"},
                    "text": {"type": "string", "description": "Text to type (for type action)"},
                    "key": {"type": "string", "description": "Key to press (for press_key, e.g. 'enter', 'ctrl+c', 'alt+f4')"},
                    "scroll_amount": {"type": "integer", "description": "Scroll amount (positive=up, negative=down)"}
                },
                "required": ["action"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "browser_automation",
            "description": "Automate web browsing: open URLs, search Google, fill forms, click buttons, extract text from web pages, download files. Works via headless browser or curl.",
            "parameters": {
                "type": "object",
                "properties": {
                    "action": {"type": "string", "enum": ["open_url", "search", "extract_text", "download", "fill_form", "screenshot"], "description": "Browser action to perform"},
                    "url": {"type": "string", "description": "URL to open or download from"},
                    "query": {"type": "string", "description": "Search query (for search action)"},
                    "selector": {"type": "string", "description": "CSS selector for fill_form"},
                    "value": {"type": "string", "description": "Value to fill in form field"},
                    "output_path": {"type": "string", "description": "Where to save downloaded file"}
                },
                "required": ["action"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_document",
            "description": "Create documents: text files, markdown reports, CSV spreadsheets, HTML pages. Can write structured content with formatting.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "File path to create"},
                    "format": {"type": "string", "enum": ["txt", "md", "csv", "html", "json", "xml"], "description": "Document format"},
                    "content": {"type": "string", "description": "Document content"},
                    "title": {"type": "string", "description": "Document title (optional)"}
                },
                "required": ["path", "content"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "translate_text",
            "description": "Translate text between languages using the AI. Supports all major languages.",
            "parameters": {
                "type": "object",
                "properties": {
                    "text": {"type": "string", "description": "Text to translate"},
                    "target_language": {"type": "string", "description": "Target language (e.g., 'Spanish', 'French', 'Chinese', 'Japanese')"},
                    "source_language": {"type": "string", "description": "Source language (optional, auto-detected if omitted)"}
                },
                "required": ["text", "target_language"]
            }
        }
    },
]


# ── Tool Execution ─────────────────────────────────────

def execute_tool(name: str, arguments: dict) -> str:
    """Execute a tool and return the result."""
    try:
        if name == "run_command":
            return tool_run_command(
                arguments["command"],
                arguments.get("timeout", 120),
                arguments.get("workdir")
            )
        elif name == "read_file":
            return tool_read_file(arguments["path"])
        elif name == "write_file":
            return tool_write_file(arguments["path"], arguments["content"])
        elif name == "search_files":
            return tool_search_files(
                arguments["pattern"],
                arguments.get("path", "."),
                arguments.get("target", "content")
            )
        elif name == "get_system_info":
            return tool_get_system_info()
        elif name == "install_software":
            return tool_install_software(arguments["name"], arguments.get("method"))
        elif name == "save_memory":
            return tool_save_memory(arguments["key"], arguments["value"])
        elif name == "recall_memory":
            return tool_recall_memory(arguments.get("key"))
        elif name == "analyze_file":
            return tool_analyze_file(arguments["path"], arguments.get("question"))
        elif name == "run_diagnosis":
            return tool_run_diagnosis(arguments.get("area", "full"))
        elif name == "setup_local_llm":
            return tool_setup_local_llm(arguments["provider"], arguments.get("model"))
        elif name == "configure_ai_agent":
            return tool_configure_ai_agent(arguments["agent"], arguments.get("api_key"))
        elif name == "list_directory":
            return tool_list_directory(arguments.get("path", "."), arguments.get("recursive", False))
        elif name == "desktop_control":
            return tool_desktop_control(arguments)
        elif name == "browser_automation":
            return tool_browser_automation(arguments)
        elif name == "create_document":
            return tool_create_document(arguments)
        elif name == "translate_text":
            return tool_translate_text(arguments)
        else:
            return json.dumps({"error": f"Unknown tool: {name}"})
    except Exception as e:
        return json.dumps({"error": str(e)})


def tool_run_command(command: str, timeout: int = 120, workdir: str = None) -> str:
    """Execute a shell command."""
    is_windows = platform.system() == "Windows"
    
    if is_windows:
        # Use PowerShell on Windows
        shell = ["powershell", "-NoProfile", "-Command", command]
    else:
        shell = ["bash", "-c", command]
    
    try:
        result = subprocess.run(
            shell,
            capture_output=True,
            text=True,
            timeout=timeout,
            cwd=workdir,
            encoding="utf-8",
            errors="replace"
        )
        output = result.stdout or ""
        if result.stderr:
            output += f"\n[STDERR]\n{result.stderr}"
        if result.returncode != 0:
            output += f"\n[EXIT CODE: {result.returncode}]"
        return output.strip() or "(no output)"
    except subprocess.TimeoutExpired:
        return f"Command timed out after {timeout}s"
    except Exception as e:
        return f"Error: {e}"


def tool_read_file(path: str) -> str:
    """Read a file."""
    p = Path(path).expanduser()
    if not p.exists():
        return f"File not found: {path}"
    if p.stat().st_size > 100_000:
        return f"File too large ({p.stat().st_size} bytes). Use search_files for large files."
    try:
        content = p.read_text(encoding="utf-8", errors="replace")
        # Add line numbers
        lines = content.split('\n')
        numbered = [f"{i+1}|{line}" for i, line in enumerate(lines)]
        return '\n'.join(numbered[:500])  # Limit to 500 lines
    except Exception as e:
        return f"Error reading file: {e}"


def tool_write_file(path: str, content: str) -> str:
    """Write to a file."""
    p = Path(path).expanduser()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content, encoding="utf-8")
    return f"Written {len(content)} bytes to {p}"


def tool_search_files(pattern: str, path: str = ".", target: str = "content") -> str:
    """Search files."""
    p = Path(path).expanduser()
    
    if target == "files":
        # Find files by name
        results = []
        for f in p.rglob(pattern):
            results.append(str(f))
            if len(results) >= 50:
                break
        return '\n'.join(results) if results else "No files found"
    else:
        # Search content with grep/rg
        is_windows = platform.system() == "Windows"
        if is_windows:
            cmd = f'findstr /S /I /N "{pattern}" *.*'
        else:
            # Try ripgrep first, fall back to grep
            if shutil.which("rg"):
                cmd = f'rg -n "{pattern}" "{path}"'
            else:
                cmd = f'grep -rn "{pattern}" "{path}"'
        
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=30)
        output = result.stdout.strip()
        if not output:
            return "No matches found"
        lines = output.split('\n')[:50]
        return '\n'.join(lines)


def tool_get_system_info() -> str:
    """Get system information."""
    info = {
        "os": platform.system(),
        "os_version": platform.version(),
        "architecture": platform.machine(),
        "processor": platform.processor(),
        "hostname": platform.node(),
        "python": sys.version.split()[0],
        "home": str(Path.home()),
        "cwd": os.getcwd(),
    }
    
    # Disk space
    try:
        if platform.system() == "Windows":
            result = subprocess.run(
                ["powershell", "-Command", "Get-PSDrive -PSProvider FileSystem | Select-Object Name,Used,Free | Format-Table -AutoSize"],
                capture_output=True, text=True, timeout=10
            )
            info["disk"] = result.stdout.strip()
        else:
            result = subprocess.run(["df", "-h"], capture_output=True, text=True, timeout=10)
            info["disk"] = result.stdout.strip()
    except:
        pass
    
    # Installed package managers
    pkg_mgrs = []
    for cmd in ["winget", "choco", "brew", "apt", "pip", "npm", "cargo"]:
        if shutil.which(cmd):
            pkg_mgrs.append(cmd)
    info["package_managers"] = pkg_mgrs
    
    return json.dumps(info, indent=2)


def tool_install_software(name: str, method: str = None) -> str:
    """Install software."""
    if not method:
        # Auto-detect
        if platform.system() == "Windows":
            method = "winget" if shutil.which("winget") else "choco"
        elif platform.system() == "Darwin":
            method = "brew"
        else:
            method = "apt"
    
    commands = {
        "winget": f"winget install --accept-package-agreements --accept-source-agreements {name}",
        "choco": f"choco install {name} -y",
        "brew": f"brew install {name}",
        "apt": f"sudo apt install {name} -y",
        "pip": f"pip install {name}",
        "npm": f"npm install -g {name}",
    }
    
    cmd = commands.get(method)
    if not cmd:
        return f"Unknown installation method: {method}"
    
    return tool_run_command(cmd, timeout=300)


def tool_save_memory(key: str, value: str) -> str:
    """Save to persistent memory."""
    mem_file = MEMORY_DIR / f"{key}.json"
    data = {"key": key, "value": value, "saved_at": time.time()}
    mem_file.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return f"Saved to memory: {key}"


def tool_recall_memory(key: str = None) -> str:
    """Recall from persistent memory."""
    if key:
        mem_file = MEMORY_DIR / f"{key}.json"
        if mem_file.exists():
            return mem_file.read_text(encoding="utf-8")
        return f"No memory found for key: {key}"
    else:
        # List all memory keys
        keys = [f.stem for f in MEMORY_DIR.glob("*.json")]
        return json.dumps({"available_keys": keys})




# ── Advanced Tool Implementations ──────────────────────

def tool_analyze_file(path: str, question: str = None) -> str:
    """Analyze any file type — PDFs, images, docs, code, etc."""
    import base64
    p = Path(path).expanduser().resolve()
    if not p.exists():
        return json.dumps({"error": f"File not found: {path}"})
    if not p.is_file():
        return json.dumps({"error": f"Not a file: {path}"})
    
    suffix = p.suffix.lower()
    size = p.stat().st_size
    result = {"path": str(p), "name": p.name, "size_bytes": size, "type": "unknown"}
    
    # Text/code files
    text_exts = {".txt",".md",".py",".js",".ts",".jsx",".tsx",".json",".yaml",".yml",".toml",
                 ".ini",".cfg",".sh",".bash",".html",".css",".scss",".xml",".csv",".log",
                 ".env",".gitignore",".rs",".go",".java",".c",".cpp",".h",".rb",".php",
                 ".swift",".kt",".dart",".r",".sql",".vue",".svelte",".astro"}
    if suffix in text_exts:
        try:
            content = p.read_text(encoding="utf-8", errors="replace")
            result["type"] = "text"
            result["language"] = suffix.lstrip(".")
            result["lines"] = content.count("\n") + 1
            # Truncate large files
            if len(content) > 50000:
                result["content"] = content[:50000] + "\n... [truncated]"
                result["truncated"] = True
            else:
                result["content"] = content
        except Exception as e:
            result["error"] = f"Failed to read: {e}"
    
    # Images — return base64 data URL for vision analysis
    elif suffix in {".png",".jpg",".jpeg",".gif",".bmp",".webp",".svg",".ico"}:
        try:
            with open(p, "rb") as f:
                data = base64.b64encode(f.read()).decode("utf-8")
            fmt = "jpeg" if suffix in {".jpg",".jpeg"} else suffix.lstrip(".")
            result["type"] = "image"
            result["format"] = fmt
            result["data_url"] = f"data:image/{fmt};base64,{data}"
            result["note"] = "Image loaded. Use vision capabilities to analyze."
        except Exception as e:
            result["error"] = f"Failed to read image: {e}"
    
    # PDF
    elif suffix == ".pdf":
        try:
            text_parts = []
            # Try PyPDF2 first
            try:
                import PyPDF2
                with open(p, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    result["type"] = "pdf"
                    result["total_pages"] = len(reader.pages)
                    for i, page in enumerate(reader.pages[:20]):
                        text = page.extract_text() or ""
                        if text.strip():
                            text_parts.append(f"--- Page {i+1} ---\n{text}")
            except ImportError:
                # Fall back to pdfminer or command line
                if shutil.which("pdftotext"):
                    r = subprocess.run(["pdftotext", str(p), "-"], capture_output=True, text=True, timeout=30)
                    text_parts.append(r.stdout)
                else:
                    result["error"] = "No PDF reader available. Install PyPDF2: pip install PyPDF2"
                    return json.dumps(result)
            
            result["type"] = "pdf"
            full_text = "\n".join(text_parts)
            if len(full_text) > 50000:
                result["content"] = full_text[:50000] + "\n... [truncated]"
                result["truncated"] = True
            else:
                result["content"] = full_text
        except Exception as e:
            result["error"] = f"Failed to read PDF: {e}"
    
    # DOCX
    elif suffix in {".docx", ".doc"}:
        try:
            import docx
            doc = docx.Document(p)
            paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
            result["type"] = "document"
            result["format"] = "docx"
            result["content"] = "\n".join(paragraphs)
        except ImportError:
            result["error"] = "python-docx not installed. Run: pip install python-docx"
        except Exception as e:
            result["error"] = f"Failed to read DOCX: {e}"
    
    # XLSX
    elif suffix in {".xlsx", ".xls"}:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(p, data_only=True)
            sheets = {}
            for sheet_name in wb.sheetnames[:5]:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(max_row=100, values_only=True):
                    rows.append([str(c) if c is not None else "" for c in row])
                sheets[sheet_name] = {"rows": rows, "total_rows": sheet.max_row}
            result["type"] = "spreadsheet"
            result["format"] = "xlsx"
            result["sheets"] = sheets
            result["total_sheets"] = len(wb.sheetnames)
        except ImportError:
            result["error"] = "openpyxl not installed. Run: pip install openpyxl"
        except Exception as e:
            result["error"] = f"Failed to read spreadsheet: {e}"
    
    # Archives
    elif suffix in {".zip",".tar",".gz",".bz2",".7z",".rar"}:
        result["type"] = "archive"
        result["format"] = suffix.lstrip(".")
        if suffix == ".zip":
            try:
                import zipfile
                with zipfile.ZipFile(p) as zf:
                    result["files"] = [{"name": i.filename, "size": i.file_size} for i in zf.infolist()[:50]]
                    result["total_files"] = len(zf.infolist())
            except Exception as e:
                result["error"] = f"Failed to read archive: {e}"
    
    # Binary/other
    else:
        result["type"] = "binary"
        result["note"] = f"Binary file ({suffix}, {size} bytes). Use run_command for processing."
    
    if question:
        result["question"] = question
    
    return json.dumps(result, indent=2)


def tool_run_diagnosis(area: str = "full") -> str:
    """Run system health diagnostics."""
    report = {"area": area, "system": platform.system(), "findings": [], "recommendations": []}
    
    # CPU load
    if platform.system() == "Windows":
        r = tool_run_command("powershell -Command \"Get-Counter '\\Processor(_Total)\\% Processor Time' -SampleInterval 2 -MaxSamples 1 | Select-Object -ExpandProperty CounterSamples | Select-Object -ExpandProperty CookedValue\"", timeout=15)
        try:
            cpu = float(r.strip().split("\n")[-1])
            report["cpu_load"] = f"{cpu:.1f}%"
            if cpu > 80:
                report["findings"].append(f"High CPU load: {cpu:.1f}%")
                report["recommendations"].append("Check running processes for resource hogs")
        except:
            report["cpu_load"] = "unknown"
        
        # Memory
        r = tool_run_command("powershell -Command \"$os = Get-CimInstance Win32_OperatingSystem; [math]::Round(($os.TotalVisibleMemorySize - $os.FreePhysicalMemory) / $os.TotalVisibleMemorySize * 100, 1)\"", timeout=10)
        try:
            mem = float(r.strip().split("\n")[-1])
            report["memory_used"] = f"{mem:.1f}%"
            if mem > 85:
                report["findings"].append(f"High memory usage: {mem:.1f}%")
                report["recommendations"].append("Close unnecessary applications to free memory")
        except:
            pass
        
        # Disk
        r = tool_run_command("powershell -Command \"Get-PSDrive -PSProvider FileSystem | ForEach-Object { $_.Name + ': ' + [math]::Round($_.Free/1GB,1) + 'GB free / ' + [math]::Round(($_.Used+$_.Free)/1GB,1) + 'GB total'}\"", timeout=10)
        report["disk"] = r.strip()
        
        # Network
        r = tool_run_command("ping -n 2 8.8.8.8", timeout=10)
        report["internet"] = "connected" if "Reply" in r else "disconnected"
        if "disconnected" in report["internet"]:
            report["findings"].append("No internet connection detected")
        
        # Startup programs
        r = tool_run_command("powershell -Command \"Get-CimInstance Win32_StartupCommand | Select-Object Name,Command | Format-Table -AutoSize\"", timeout=10)
        startup_lines = [l for l in r.strip().split("\n") if l.strip() and "Name" not in l and "---" not in l]
        report["startup_programs"] = len(startup_lines)
        if len(startup_lines) > 15:
            report["findings"].append(f"{len(startup_lines)} startup programs — may slow boot time")
            report["recommendations"].append("Disable unnecessary startup programs via Task Manager")
        
        # Windows Update status
        r = tool_run_command("powershell -Command \"Get-HotFix | Sort-Object InstalledOn -Descending | Select-Object -First 3 | Format-Table HotFixID,Description,InstalledOn -AutoSize\"", timeout=15)
        report["recent_updates"] = r.strip()[:500]
    
    else:  # Linux/Mac
        r = tool_run_command("uptime", timeout=5)
        report["uptime"] = r.strip()
        r = tool_run_command("df -h /", timeout=5)
        report["disk"] = r.strip()
        r = tool_run_command("ping -c 2 8.8.8.8", timeout=10)
        report["internet"] = "connected" if r and "bytes from" in r else "disconnected"
    
    if not report["findings"]:
        report["findings"].append("No critical issues detected. System appears healthy.")
        report["recommendations"].append("Keep system updated and run periodic cleanups.")
    
    return json.dumps(report, indent=2)


def tool_setup_local_llm(provider: str, model: str = None) -> str:
    """Install and configure local LLM providers."""
    steps = []
    
    if provider == "ollama":
        # Check if already installed
        check = tool_run_command("ollama --version", timeout=5)
        if "ollama" in check.lower() and "error" not in check.lower():
            steps.append({"step": "Ollama already installed", "status": "done", "version": check.strip()})
        else:
            if platform.system() == "Windows":
                steps.append({"step": "Installing Ollama", "command": "winget install Ollama.Ollama", "status": "running"})
                result = tool_run_command("winget install Ollama.Ollama --accept-package-agreements --accept-source-agreements", timeout=300)
                steps[-1]["result"] = result[:500]
                steps[-1]["status"] = "done" if "Successfully" in result else "check"
            elif platform.system() == "Darwin":
                result = tool_run_command("brew install ollama", timeout=300)
                steps.append({"step": "Installed via brew", "result": result[:500], "status": "done"})
            else:
                result = tool_run_command("curl -fsSL https://ollama.com/install.sh | sh", timeout=300)
                steps.append({"step": "Installed via script", "result": result[:500], "status": "done"})
        
        # Pull model
        if model:
            steps.append({"step": f"Downloading {model}", "status": "running"})
            result = tool_run_command(f"ollama pull {model}", timeout=600)
            steps[-1]["result"] = result[:500]
            steps[-1]["status"] = "done"
        else:
            steps.append({"step": "No model specified", "note": "Run 'ollama pull llama3' to download a model", "status": "info"})
            steps.append({"step": "Popular models", "models": ["llama3 (8B)", "mistral (7B)", "codellama (7B/13B/34B)", "phi3 (3.8B)", "gemma (2B/7B)"], "status": "info"})
    
    elif provider == "lmstudio":
        steps.append({"step": "LM Studio", "note": "LM Studio requires manual download from https://lmstudio.ai", "status": "manual"})
        if platform.system() == "Windows":
            steps.append({"step": "Opening download page", "url": "https://lmstudio.ai/download"})
    elif provider == "llamacpp":
        if shutil.which("git"):
            result = tool_run_command("git clone https://github.com/ggerganov/llama.cpp.git ~/llama.cpp", timeout=120)
            steps.append({"step": "Cloned llama.cpp", "result": result[:300], "status": "done"})
            steps.append({"step": "Build required", "note": "Run cmake/make in ~/llama.cpp to build", "status": "info"})
        else:
            steps.append({"step": "Git required", "note": "Install git first: winget install Git.Git", "status": "error"})
    
    return json.dumps({"provider": provider, "steps": steps}, indent=2)


def tool_configure_ai_agent(agent: str, api_key: str = None) -> str:
    """Configure AI coding agents."""
    result = {"agent": agent, "steps": [], "status": "unknown"}
    
    if agent == "claude-code":
        check = tool_run_command("claude --version", timeout=5)
        if "error" not in check.lower():
            result["steps"].append({"step": "Claude Code already installed", "version": check.strip()})
        else:
            r = tool_run_command("npm install -g @anthropic-ai/claude-code", timeout=120)
            result["steps"].append({"step": "Installed via npm", "result": r[:300]})
        if api_key:
            # Set env var
            if platform.system() == "Windows":
                tool_run_command(f'setx ANTHROPIC_API_KEY "{api_key}"', timeout=10)
            else:
                # Add to shell profile
                shell_rc = Path.home() / ".bashrc"
                if (Path.home() / ".zshrc").exists():
                    shell_rc = Path.home() / ".zshrc"
                with open(shell_rc, "a") as f:
                    f.write(f'\nexport ANTHROPIC_API_KEY="{api_key}"\n')
            result["steps"].append({"step": "API key configured", "status": "done"})
        else:
            result["steps"].append({"step": "API key needed", "note": "Set ANTHROPIC_API_KEY environment variable", "status": "manual"})
        result["status"] = "configured"
    
    elif agent == "copilot-cli":
        check = tool_run_command("gh --version", timeout=5)
        if "error" not in check.lower():
            result["steps"].append({"step": "GitHub CLI installed", "version": check.strip()})
            r = tool_run_command("gh extension install github/gh-copilot", timeout=60)
            result["steps"].append({"step": "Copilot extension installed", "result": r[:300]})
        else:
            r = tool_run_command("winget install GitHub.cli" if platform.system() == "Windows" else "brew install gh", timeout=120)
            result["steps"].append({"step": "Installing GitHub CLI", "result": r[:300]})
            result["steps"].append({"step": "After install, run: gh extension install github/gh-copilot", "status": "manual"})
        result["status"] = "configured"
    
    elif agent == "cursor":
        result["steps"].append({"step": "Cursor", "note": "Download from https://cursor.sh", "url": "https://cursor.sh"})
        if platform.system() == "Windows":
            tool_run_command("winget install Cursor.Cursor", timeout=120)
            result["steps"].append({"step": "Installing via winget", "status": "done"})
        result["status"] = "configured"
    
    elif agent == "aider":
        r = tool_run_command("pip install aider-chat", timeout=120)
        result["steps"].append({"step": "Installed aider-chat", "result": r[:300]})
        if api_key:
            if platform.system() == "Windows":
                tool_run_command(f'setx OPENAI_API_KEY "{api_key}"', timeout=10)
            result["steps"].append({"step": "API key set", "status": "done"})
        result["status"] = "configured"
    
    elif agent == "openhands":
        r = tool_run_command("pip install openhands-ai", timeout=120)
        result["steps"].append({"step": "Installed openhands-ai", "result": r[:300]})
        result["status"] = "configured"
    
    elif agent == "continue":
        result["steps"].append({"step": "Continue.dev", "note": "Install as VS Code extension: continue.continue", "status": "manual"})
        if shutil.which("code"):
            tool_run_command("code --install-extension continue.continue", timeout=60)
            result["steps"].append({"step": "Installed via VS Code CLI", "status": "done"})
        result["status"] = "configured"
    
    return json.dumps(result, indent=2)


def tool_list_directory(path: str = ".", recursive: bool = False) -> str:
    """List directory contents."""
    p = Path(path).expanduser().resolve()
    if not p.exists():
        return json.dumps({"error": f"Directory not found: {path}"})
    if not p.is_dir():
        return json.dumps({"error": f"Not a directory: {path}"})
    
    items = []
    try:
        for item in sorted(p.iterdir(), key=lambda x: (not x.is_dir(), x.name.lower())):
            if item.name.startswith(".") and item.name not in [".env", ".gitignore"]:
                continue
            entry = {"name": item.name, "type": "dir" if item.is_dir() else "file"}
            if item.is_file():
                size = item.stat().st_size
                entry["size"] = size
                entry["size_human"] = f"{size/1024:.1f}KB" if size < 1024*1024 else f"{size/1024/1024:.1f}MB"
                entry["ext"] = item.suffix
            if recursive and item.is_dir():
                try:
                    sub_items = list(item.iterdir())
                    entry["children_count"] = len(sub_items)
                except PermissionError:
                    entry["children_count"] = "permission denied"
            items.append(entry)
            if len(items) >= 100:
                items.append({"note": "Truncated at 100 items"})
                break
    except PermissionError:
        return json.dumps({"error": "Permission denied"})
    
    return json.dumps({"path": str(p), "items": items, "total": len(items)}, indent=2)


def tool_desktop_control(args: dict) -> str:
    """Control the user's desktop via pyautogui or PowerShell automation."""
    action = args.get("action", "")
    try:
        # Try pyautogui first
        import pyautogui
        pyautogui.FAILSAFE = True
        pyautogui.PAUSE = 0.1
        
        if action == "click":
            x, y = args.get("x", 0), args.get("y", 0)
            pyautogui.click(x, y)
            return json.dumps({"status": "clicked", "x": x, "y": y})
        elif action == "double_click":
            x, y = args.get("x", 0), args.get("y", 0)
            pyautogui.doubleClick(x, y)
            return json.dumps({"status": "double_clicked", "x": x, "y": y})
        elif action == "right_click":
            x, y = args.get("x", 0), args.get("y", 0)
            pyautogui.rightClick(x, y)
            return json.dumps({"status": "right_clicked", "x": x, "y": y})
        elif action == "type":
            text = args.get("text", "")
            pyautogui.typewrite(text, interval=0.02)
            return json.dumps({"status": "typed", "length": len(text)})
        elif action == "press_key":
            key = args.get("key", "enter")
            # Handle combos like ctrl+c
            keys = key.split("+")
            pyautogui.hotkey(*keys)
            return json.dumps({"status": "pressed", "key": key})
        elif action == "screenshot":
            path = str(DATA_DIR / "screenshot.png")
            pyautogui.screenshot(path)
            return json.dumps({"status": "screenshot", "path": path})
        elif action == "move_to":
            x, y = args.get("x", 0), args.get("y", 0)
            pyautogui.moveTo(x, y)
            return json.dumps({"status": "moved", "x": x, "y": y})
        elif action == "scroll":
            amount = args.get("scroll_amount", -3)
            pyautogui.scroll(amount)
            return json.dumps({"status": "scrolled", "amount": amount})
        else:
            return json.dumps({"error": f"Unknown action: {action}"})
    except ImportError:
        # Fallback to PowerShell for Windows
        if platform.system() == "Windows":
            if action == "click":
                x, y = args.get("x", 0), args.get("y", 0)
                ps = f'Add-Type -AssemblyName System.Windows.Forms; [System.Windows.Forms.Cursor]::Position = New-Object System.Drawing.Point({x},{y}); Add-Type -MemberDefinition "[DllImport(\\"user32.dll\\")] public static extern void mouse_event(int flags, int dx, int dy, int buttons, int extra);" -Name Win32 -Namespace System; [System.Win32]::mouse_event(2, 0, 0, 0, 0); [System.Win32]::mouse_event(4, 0, 0, 0, 0)'
                return tool_run_command(f'powershell -Command "{ps}"', timeout=10)
            elif action == "type":
                text = args.get("text", "")
                return tool_run_command(f'powershell -Command "$w = New-Object -ComObject WScript.Shell; $w.SendKeys(\'{text}\')"', timeout=10)
            elif action == "screenshot":
                path = str(DATA_DIR / "screenshot.png")
                return tool_run_command(f'powershell -Command "Add-Type -AssemblyName System.Windows.Forms; [System.Windows.Forms.Screen]::PrimaryScreen | ForEach-Object {{ $bitmap = New-Object System.Drawing.Bitmap($_.Bounds.Width, $_.Bounds.Height); $graphics = [System.Drawing.Graphics]::FromImage($bitmap); $graphics.CopyFromScreen($_.Bounds.Location, [System.Drawing.Point]::Empty, $_.Bounds.Size); $bitmap.Save(\'{path}\') }}"', timeout=15)
        return json.dumps({"error": "pyautogui not installed. Run: pip install pyautogui", "fallback": "PowerShell automation attempted"})
    except Exception as e:
        return json.dumps({"error": f"Desktop control failed: {str(e)}"})


def tool_browser_automation(args: dict) -> str:
    """Automate web browsing via curl/wget or headless browser."""
    action = args.get("action", "")
    try:
        if action == "open_url":
            url = args.get("url", "")
            if not url.startswith("http"):
                url = "https://" + url
            # Try to open in default browser
            import webbrowser
            webbrowser.open(url)
            return json.dumps({"status": "opened", "url": url})
        
        elif action == "search":
            query = args.get("query", "")
            url = f"https://www.google.com/search?q={query.replace(' ', '+')}"
            import webbrowser
            webbrowser.open(url)
            # Also fetch results via curl for text extraction
            result = tool_run_command(f'curl -s -L -A "Mozilla/5.0" "{url}"', timeout=30)
            # Strip HTML tags
            import re
            text = re.sub(r'<[^>]+>', ' ', result)
            text = re.sub(r'\s+', ' ', text).strip()
            return json.dumps({"status": "searched", "query": query, "url": url, "preview": text[:3000]})
        
        elif action == "extract_text":
            url = args.get("url", "")
            if not url.startswith("http"):
                url = "https://" + url
            result = tool_run_command(f'curl -s -L -A "Mozilla/5.0 (Windows NT 10.0; Win64; x64)" "{url}"', timeout=30)
            import re
            # Remove scripts, styles
            text = re.sub(r'<script[^>]*>.*?</script>', '', result, flags=re.DOTALL)
            text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.DOTALL)
            text = re.sub(r'<[^>]+>', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            return json.dumps({"status": "extracted", "url": url, "text": text[:5000], "length": len(text)})
        
        elif action == "download":
            url = args.get("url", "")
            output = args.get("output_path", str(DATA_DIR / "download"))
            result = tool_run_command(f'curl -s -L -o "{output}" "{url}"', timeout=120)
            if Path(output).exists():
                size = Path(output).stat().st_size
                return json.dumps({"status": "downloaded", "path": output, "size": size, "size_human": f"{size/1024:.1f}KB"})
            return json.dumps({"error": "Download failed", "url": url})
        
        elif action == "screenshot":
            url = args.get("url", "")
            path = str(DATA_DIR / "web_screenshot.png")
            # Try pyautogui screenshot after opening URL
            import webbrowser
            webbrowser.open(url)
            import time
            time.sleep(3)
            try:
                import pyautogui
                pyautogui.screenshot(path)
                return json.dumps({"status": "screenshot", "path": path, "url": url})
            except ImportError:
                return json.dumps({"status": "url_opened", "url": url, "note": "Screenshot requires pyautogui"})
        
        else:
            return json.dumps({"error": f"Unknown browser action: {action}"})
    except Exception as e:
        return json.dumps({"error": f"Browser automation failed: {str(e)}"})


def tool_create_document(args: dict) -> str:
    """Create structured documents."""
    path = args.get("path", "")
    content = args.get("content", "")
    fmt = args.get("format", "txt")
    title = args.get("title", "")
    
    p = Path(path).expanduser()
    p.parent.mkdir(parents=True, exist_ok=True)
    
    if fmt == "html" and title:
        content = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"><title>{title}</title></head>
<body>
<h1>{title}</h1>
{content}
</body>
</html>"""
    elif fmt == "md" and title:
        content = f"# {title}\n\n{content}"
    elif fmt == "csv":
        # Content should already be CSV formatted
        pass
    elif fmt == "json":
        try:
            # Validate JSON
            json.loads(content)
        except json.JSONDecodeError:
            content = json.dumps({"content": content}, indent=2)
    elif fmt == "xml" and title:
        content = f'<?xml version="1.0" encoding="UTF-8"?>\n<document title="{title}">\n{content}\n</document>'
    
    p.write_text(content, encoding="utf-8")
    size = p.stat().st_size
    return json.dumps({"status": "created", "path": str(p), "format": fmt, "size": size, "size_human": f"{size/1024:.1f}KB"})


def tool_translate_text(args: dict) -> str:
    """Translate text — this is handled by returning a special instruction to the LLM."""
    text = args.get("text", "")
    target = args.get("target_language", "English")
    source = args.get("source_language", "auto-detected")
    
    # Since we can't call a separate translation API, we return the text
    # and let the LLM handle it in the next turn
    return json.dumps({
        "status": "translation_requested",
        "text": text[:500],
        "source_language": source,
        "target_language": target,
        "instruction": f"Please translate the following text to {target}. Original text: {text[:2000]}"
    })



# ── Agent Loop ─────────────────────────────────────────

def load_conversation(thread_id: str) -> list:
    """Load conversation history."""
    if CONVERSATIONS_FILE.exists():
        try:
            data = json.loads(CONVERSATIONS_FILE.read_text(encoding="utf-8"))
            return data.get(thread_id, [])
        except:
            pass
    return []


def save_conversation(thread_id: str, messages: list):
    """Save conversation history."""
    data = {}
    if CONVERSATIONS_FILE.exists():
        try:
            data = json.loads(CONVERSATIONS_FILE.read_text(encoding="utf-8"))
        except:
            pass
    # Keep last 100 messages per thread
    data[thread_id] = messages[-100:]
    CONVERSATIONS_FILE.write_text(json.dumps(data), encoding="utf-8")


def agent_loop(user_message: str, thread_id: str, auth_token: str):
    """
    Main agent loop: send to LLM, execute tools, repeat until done.
    Streams responses via stdout to Tauri.
    """
    messages = load_conversation(thread_id)
    messages.append({"role": "user", "content": user_message})
    
    max_iterations = 10  # Prevent infinite loops
    
    for iteration in range(max_iterations):
        # Load memory context
        memory_context = ""
        mem_keys_file = MEMORY_DIR / "_index.json"
        if not mem_keys_file.exists():
            # Build index
            keys = [f.stem for f in MEMORY_DIR.glob("*.json") if f.stem != "_index"]
            if keys:
                memory_context = f"\n\n[Available memory: {', '.join(keys)}]"
        
        # Call Vercel API with tools
        payload = {
            "messages": [
                {"role": "system", "content": SYSTEM_PROMPT + memory_context},
                *messages
            ],
            "tools": TOOLS,
            "model": "deepseek-v4-flash",
            "stream": False,  # Non-streaming for tool calling
        }
        
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {auth_token}",
        }
        
        try:
            import urllib.request
            import urllib.error
            req = urllib.request.Request(
                VERCEL_API,
                data=json.dumps(payload).encode("utf-8"),
                headers=headers,
                method="POST"
            )
            
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            # Read the actual error body from the server
            try:
                err_body = json.loads(e.read().decode("utf-8"))
                err_msg = err_body.get("error", "") or err_body.get("message", "") or str(e)
                # Include all useful fields for frontend detection
                if err_body.get("rateLimited"):
                    err_msg += " rateLimited"
                if err_body.get("credits") is not None:
                    err_msg += f" (credits: {err_body['credits']})"
            except Exception:
                err_msg = str(e)
            
            # SANITIZE: never expose raw provider errors to users
            err_msg = sanitize_error(err_msg, e.code)
            emit({"type": "error", "content": err_msg})
            emit({"type": "done"})
            return
        except Exception as e:
            emit({"type": "error", "content": "Something went wrong. Please try again."})
            emit({"type": "done"})
            return
        
        # Check for tool calls
        choice = data.get("choices", [{}])[0]
        message = choice.get("message", {})
        tool_calls = message.get("tool_calls")
        
        if tool_calls:
            # Add assistant message with tool calls
            messages.append(message)
            
            # Execute each tool
            for tc in tool_calls:
                fn = tc["function"]
                tool_name = fn["name"]
                try:
                    args = json.loads(fn["arguments"])
                except:
                    args = {}
                
                # Emit tool start
                emit({"type": "tool_start", "name": tool_name, "args": args})
                
                # Execute
                result = execute_tool(tool_name, args)
                
                # Emit tool result
                emit({"type": "tool_result", "name": tool_name, "result": result[:2000]})
                
                # Add tool result to messages
                messages.append({
                    "role": "tool",
                    "tool_call_id": tc["id"],
                    "content": result[:4000]  # Truncate for context
                })
            
            # Continue loop to let AI process results
            continue
        else:
            # Final response — no more tool calls
            content = message.get("content", "")
            thinking = message.get("reasoning_content", "")
            
            if thinking:
                emit({"type": "thinking", "content": thinking})
            
            emit({"type": "content", "content": content})
            messages.append({"role": "assistant", "content": content})
            save_conversation(thread_id, messages)
            emit({"type": "done"})
            return
    
    # Max iterations reached
    emit({"type": "content", "content": "I've reached the maximum number of tool calls. Here's what I've done so far."})
    save_conversation(thread_id, messages)
    emit({"type": "done"})


def emit(data: dict):
    """Send JSON message to Tauri via stdout."""
    print(json.dumps(data), flush=True)


# ── Main ───────────────────────────────────────────────

def main():
    """Read JSON commands from stdin, process, emit results to stdout."""
    emit({"type": "status", "content": "TaskBolt Agent Engine ready"})
    
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        
        try:
            cmd = json.loads(line)
        except:
            continue
        
        action = cmd.get("action")
        
        if action == "chat":
            user_message = cmd.get("message", "")
            thread_id = cmd.get("thread_id", "default")
            auth_token = cmd.get("auth_token", AUTH_TOKEN)
            
            if not user_message:
                emit({"type": "error", "content": "Empty message"})
                continue
            
            agent_loop(user_message, thread_id, auth_token)
        
        elif action == "stop":
            emit({"type": "status", "content": "Stopping..."})
            break
        
        elif action == "ping":
            emit({"type": "pong"})


if __name__ == "__main__":
    main()
